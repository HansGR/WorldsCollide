from openpyxl import load_workbook

doors_WOB_WOR = {
    113 : 145,
    114 : 146,
    115 : 147,
    1139 : 1147,
    1137 : 1145,
    1138 : 1146,
    1140 : 1148,
    1141 : 1149,
    1142 : 1150,
    169 : 148,
    170 : 149,
    171 : 150,
    172 : 151,
    173 : 152,
    174 : 153,
    178 : 156,
    179 : 157,
    1155 : 1151,
    180 : 158
}

# Collect data
fn = 'C:\\Users\\HRIN\\Documents\\GitHub\\WorldsCollide\\LocationRandomizer-WC.xlsm'

wb = load_workbook(fn)
ws = wb['Exits']

doorID = []
door_descr = dict()
door_pair = dict()

for i in range(3,ws.max_row):
    dID = ws['A'+str(i)].value
    doorID.append(dID)
    door_descr[dID] = ws['Q'+str(i)].value
    door_pair[dID] = ws['R' + str(i)].value

# Write exit_data
f = open("map_exit_extra.py", "w")
f.write('#exit number : [doorpair ID, description]')
f.write('\nexit_data = { ')
for d in doorID:
    if d is not None:
        f.write('\n\t' + str(d) + ' : [' + str(door_pair[d]) + ', "' + str(door_descr[d]) + '"]')
        if d != doorID[-1]:
            f.write(',')
f.write('\n}\n')

# Write doors_WOB_WOR
f.write('\n# exit number in WOB : equivalent exit in WOR')
f.write('\ndoors_WOB_WOR = { ')
writedoors = [d for d in doors_WOB_WOR.keys()]
for d in writedoors:
    f.write('\n\t' + str(d) + ' : ' + str(doors_WOB_WOR[d]) )
    if d != writedoors[-1]:
        f.write(',')
f.write('\n}\n')

f.close()
